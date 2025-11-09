# app/services/importers.py
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

try:
    # openpyxl is optional; we import lazily and handle ImportError nicely
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - envs without excel support
    load_workbook = None  # type: ignore


# ---- Public API -------------------------------------------------------------

def parse_items_file(
    file_name: str,
    file_bytes: bytes,
    mime_type: Optional[str] = None,
) -> List[Tuple[str, str]]:
    """Parse a CSV or Excel file with columns (question, answer).

    Returns a list of (question, answer) in the original order.
    Header names are matched case-insensitively with Russian/English aliases:
      - question: ["question", "вопрос", "q", "term", "front"]
      - answer:   ["answer", "ответ", "a", "definition", "back"]

    Duplicates are removed by normalized question (first occurrence wins).

    Raises:
        ValueError: if file is not recognized or columns are missing.
    """
    kind = _detect_kind(file_name, mime_type, file_bytes)

    if kind == "xlsx":
        rows = _read_xlsx(file_bytes)
    else:
        rows = _read_csv(file_bytes)

    header_map = _build_header_map(rows.headers)

    q_key = header_map.get("question")
    a_key = header_map.get("answer")

    if not q_key or not a_key:
        raise ValueError(
            "Ожидались колонки 'question' и 'answer' (или их синонимы). Найдены: " + ", ".join(rows.headers)
        )

    seen: set[str] = set()
    pairs: List[Tuple[str, str]] = []
    for row in rows.iter_dicts():
        q_raw = (row.get(q_key) or "").strip()
        a_raw = (row.get(a_key) or "").strip()
        if not q_raw or not a_raw:
            continue
        q = _normalize_text(q_raw)
        a = _normalize_text(a_raw)
        dedup_key = _dedup_key(q)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        pairs.append((q, a))

    if not pairs:
        raise ValueError("Файл прочитан, но валидных пар 'вопрос-ответ' не найдено.")

    return pairs


def parse_collections_file(
    file_name: str,
    file_bytes: bytes,
    mime_type: Optional[str] = None,
) -> Dict[str, List[Tuple[str, str]]]:
    """Parse CSV/Excel with columns (title, question, answer).

    Returns mapping: {collection_title -> list[(question, answer), ...]}.
    Duplicates inside the same collection are removed by normalized question.

    Raises:
        ValueError: if required columns are missing or file is empty.
    """
    kind = _detect_kind(file_name, mime_type, file_bytes)

    if kind == "xlsx":
        rows = _read_xlsx(file_bytes)
    else:
        rows = _read_csv(file_bytes)

    header_map = _build_header_map(rows.headers)

    t_key = header_map.get("title")
    q_key = header_map.get("question")
    a_key = header_map.get("answer")
    if not t_key or not q_key or not a_key:
        raise ValueError(
            "Ожидались колонки 'title', 'question', 'answer' (или их синонимы). Найдены: " + ", ".join(rows.headers)
        )

    grouped: Dict[str, List[Tuple[str, str]]] = {}
    seen_per_title: Dict[str, set[str]] = {}

    for row in rows.iter_dicts():
        t_raw = (row.get(t_key) or "").strip()
        q_raw = (row.get(q_key) or "").strip()
        a_raw = (row.get(a_key) or "").strip()
        if not t_raw or not q_raw or not a_raw:
            continue
        title = _normalize_text(t_raw)
        q = _normalize_text(q_raw)
        a = _normalize_text(a_raw)

        dd = _dedup_key(q)
        seen = seen_per_title.setdefault(title, set())
        if dd in seen:
            continue
        seen.add(dd)
        grouped.setdefault(title, []).append((q, a))

    if not grouped:
        raise ValueError("Файл прочитан, но коллекции не обнаружены.")

    return grouped


# ---- Helpers ----------------------------------------------------------------


@dataclass(slots=True)
class _RowSet:
    headers: List[str]
    rows: List[List[str]]

    def iter_dicts(self) -> Iterable[Dict[str, str]]:
        idx = {i: h for i, h in enumerate(self.headers)}
        for r in self.rows:
            d: Dict[str, str] = {}
            for i, v in enumerate(r[: len(self.headers)]):
                # ensure str
                d[idx[i]] = v if isinstance(v, str) else ("" if v is None else str(v))
            yield d


def _build_header_map(headers: Iterable[str]) -> Dict[str, str]:
    """Build mapping canonical_key -> real_header from a header row.

    Supports Russian/English aliases and case-insensitive comparison.
    """
    aliases = {
        "title": {"title", "название", "коллекция", "collection", "deck"},
        "question": {"question", "вопрос", "q", "term", "front"},
        "answer": {"answer", "ответ", "a", "definition", "back"},
    }
    lower = {h.lower().strip(): h for h in headers}
    mapping: Dict[str, str] = {}
    for canon, names in aliases.items():
        for name in names:
            if name in lower:
                mapping[canon] = lower[name]
                break
    return mapping


def _normalize_text(s: str) -> str:
    # collapse whitespace and strip
    return " ".join((s or "").replace("\xa0", " ").split()).strip()


def _dedup_key(s: str) -> str:
    return _normalize_text(s).lower()


def _detect_kind(file_name: str, mime_type: Optional[str], file_bytes: bytes) -> str:
    name = (file_name or "").lower()
    if name.endswith(".xlsx") or (mime_type and mime_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",)):
        if load_workbook is None:
            raise ValueError("Поддержка .xlsx не установлена (нет openpyxl). Установите 'openpyxl' или пришлите CSV.")
        return "xlsx"
    # Detect by BOM or typical CSV text
    txt = _safe_peek_text(file_bytes, limit=2048)
    if "," in txt or ";" in txt or "\t" in txt or "question" in txt.lower():
        return "csv"
    if name.endswith(".csv"):
        return "csv"
    # Fall back to CSV
    return "csv"


def _safe_peek_text(data: bytes, limit: int = 1024) -> str:
    return _decode_bytes(data[:limit])


def _read_csv(file_bytes: bytes) -> _RowSet:
    text = _decode_bytes(file_bytes)
    # Try to detect dialect; if fails, assume comma
    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(text.splitlines()[0] if text.splitlines() else text, delimiters=",;\t")
    except Exception:
        class _D(csv.Dialect):
            delimiter = ','
            quotechar = '"'
            doublequote = True
            skipinitialspace = True
            lineterminator = '\n'
            quoting = csv.QUOTE_MINIMAL
        dialect = _D()
    reader = csv.reader(io.StringIO(text), dialect)
    rows: List[List[str]] = [list(r) for r in reader if any((c or '').strip() for c in r)]
    if not rows:
        return _RowSet(headers=[], rows=[])
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    # pad/truncate rows to header length
    norm_rows: List[List[str]] = []
    for r in data:
        if len(r) < len(headers):
            r = r + ["" for _ in range(len(headers) - len(r))]
        norm_rows.append(r[: len(headers)])
    return _RowSet(headers=headers, rows=norm_rows)


def _read_xlsx(file_bytes: bytes) -> _RowSet:
    if load_workbook is None:
        raise ValueError("Поддержка .xlsx не установлена (нет openpyxl). Установите 'openpyxl' или пришлите CSV.")
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    headers: List[str] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        vals = [(v if isinstance(v, str) else ("" if v is None else str(v))).strip() for v in row]
        if first:
            headers = [v for v in vals]
            first = False
            continue
        if any(v.strip() for v in vals):
            rows.append(vals[: len(headers)])
    wb.close()
    return _RowSet(headers=headers, rows=rows)


def _decode_bytes(b: bytes) -> str:
    # Try UTF-8 with BOM first, then UTF-8, then common Cyrillic codepages, then latin-1 as last resort.
    for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "iso-8859-5", "koi8-r", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    # If everything fails, decode as utf-8 with replacement to avoid exceptions.
    return b.decode("utf-8", errors="replace")
